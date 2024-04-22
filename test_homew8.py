import os
import shutil
import zipfile
import csv
from fileinput import filename

from PyPDF2 import PdfReader
from openpyxl import load_workbook

# Создаем папку и архив
os.mkdir('resources/tmp')
myzip = zipfile.ZipFile('resources/tmp/my_zip.zip', 'w')


# Добавляем файлы в архив
def test_archive_zip():
    myzip = zipfile.ZipFile(filename('resources/tmp/my_zip.zip'), 'w')
    myzip.write(filename('resources/username.csv'))
    myzip.write(filename('resources/file_example_XLSX_50.xlsx'))
    myzip.write(filename('resources/docs-pytest-org-en-latest.pdf'))
    myzip.close()

# Проверка файла
def test_read_zip():
    for file_info in myzip.infolist():
        print(file_info.filename, file_info.date_time, file_info.file_size)
    myzip.close()


# Разархивируем файл
def test_unzip_files():
    unzip_file = zipfile.ZipFile('resources/tmp/my_zip.zip')
    unzip_file.extractall('resources/tmp/')
    unzip_file.close()
    


# Проверка pdf
def test_read_pdf():
    reader = PdfReader('resources/tmp/docs-pytest-org-en-latest.pdf')
    page = reader.pages[0]
    text = page.extract_text()
    assert 'pytest Documentation' in text


# Проверка xlsx
def test_read_xlsx():
    workbook = load_workbook('resources/tmp/file_example_XLSX_50.xlsx')
    sheet = workbook.active
    name = sheet.cell(row=2, column=1).value
    assert 'Dulce' == name


# Проверка csv
def test_read_csv_():
    with open('resources/tmp/username.csv') as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert 'First_name' in str(headers)


# Удаляем папку
def test_remove_folder():
    shutil.rmtree('resources/tmphg/')
